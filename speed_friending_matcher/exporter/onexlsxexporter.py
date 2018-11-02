# coding=utf-8
from .exporter import Exporter
from ..core.person import MatchingFlags
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.utils import get_column_letter


class OneXlsxExporter(Exporter):
    def __init__(self, filename):
        Exporter.__init__(self)
        self._filename = filename

    def export_data(self, data):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Contacts'

        self._create_styles(workbook)
        self._write_header(worksheet, len(data))

        for person in data:
            self._write_person(worksheet, person)

        workbook.save(self._filename)

    def _create_styles(self, workbook):
        style = NamedStyle(name="title")
        style.font = Font(bold=True)
        style.fill = PatternFill(fill_type=FILL_SOLID, start_color='FFFF9900')
        workbook.add_named_style(style)
        style = NamedStyle(name="id")
        style.font = Font(bold=True)
        style.fill = PatternFill(fill_type=FILL_SOLID, start_color='FFFFFF00')
        workbook.add_named_style(style)
        style = NamedStyle(name="anonymous")
        style.fill = PatternFill(fill_type=FILL_SOLID, start_color='FFFF9999')
        workbook.add_named_style(style)

    def _write_header(self, worksheet, num_particpants):
        elements = ['#', 'Name', 'Email', 'Phone', 'Contact']
        for i, element in enumerate(elements):
            cell = worksheet[self._cell(0, i)]
            cell.value = element
            cell.style = 'title'
        for i in range(num_particpants):
            column = len(elements) + i
            cell = worksheet[self._cell(0, column)]
            cell.value = i + 1
            cell.style = 'title'

        worksheet.column_dimensions["A"].width = 5
        worksheet.column_dimensions["B"].width = 20
        worksheet.column_dimensions["C"].width = 20
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 10
        for i in range(num_particpants):
            column = get_column_letter(len(elements) + i + 1)
            worksheet.column_dimensions[column].width = 3

    def _write_person(self, worksheet, person):
        row = person.number
        cell = worksheet[self._cell(row, 0)]
        cell.value = person.number
        cell.style = 'id'
        if person.flags == MatchingFlags.match_all:
            worksheet[self._cell(row, 1)] = person.name
            worksheet[self._cell(row, 2)] = person.email
            worksheet[self._cell(row, 3)] = person.phone
        else:
            cell = worksheet[self._cell(row, 1)]
            cell.value = 'Anonymous'
            cell.style = 'anonymous'
            cell = worksheet[self._cell(row, 2)]
            cell.value = ''
            cell.style = 'anonymous'
            cell = worksheet[self._cell(row, 3)]
            cell.value = ''
            cell.style = 'anonymous'
        worksheet[self._cell(row, 4)] = self._matching_flags_to_string(person.flags)
        for other_person in person.results.marked_by_me:
            worksheet[self._cell(row, 4 + other_person.number)] = 'x'

    def _matching_flags_to_string(self, flags):
        if flags == MatchingFlags.match_all:
            return 'all'
        else:
            return 'not all'

    def _cell(self, row, column):
        return '%c%i' % (get_column_letter(column + 1), row + 1)


exporter = OneXlsxExporter
